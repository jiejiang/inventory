#!/usr/bin/env python
# -*- coding: utf-8 -*-
import cStringIO, json
import re, math, random
import sys
import os
import shutil
import zipfile
import datetime
from cStringIO import StringIO
from optparse import OptionParser
from jinja2 import Environment, FileSystemLoader
import pandas as pd
import barcode
from barcode.writer import ImageWriter
from weasyprint import HTML, CSS
from wand.image import Image
from wand.color import Color
from PyPDF2 import PdfFileMerger, PdfFileReader, PdfFileWriter
from sqlalchemy import desc, asc, Index, UniqueConstraint, and_
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from flask import current_app
from pdf2image import convert_from_path, convert_from_bytes
from faker import Faker

from ..models import City, Order, ProductInfo
from .. import db
from ..util import time_to_filename

Code128 = barcode.get_barcode_class('code128')

PROVINCE_INFO_MAP = {
    u"湖南": {'ticket_initial': 9, 'package_type': u"快递包裹"},
    u"广西": {'ticket_initial': 9, 'package_type': u"快递包裹"},
    u"海南": {'ticket_initial': 9, 'package_type': u"快递包裹"},
    u"江西": {'ticket_initial': 9, 'package_type': u"快递包裹"},
    u"福建": {'ticket_initial': 9, 'package_type': u"快递包裹"},
    u"湖北": {'ticket_initial': 9, 'package_type': u"快递包裹"},
    u"江苏": {'ticket_initial': 9, 'package_type': u"快递包裹"},
    u"上海": {'ticket_initial': 9, 'package_type': u"快递包裹"},
    u"浙江": {'ticket_initial': 9, 'package_type': u"快递包裹"},
    u"河北": {'ticket_initial': 9, 'package_type': u"快递包裹"},
    u"安徽": {'ticket_initial': 9, 'package_type': u"快递包裹"},
    u"河南": {'ticket_initial': 9, 'package_type': u"快递包裹"},
    u"山东": {'ticket_initial': 9, 'package_type': u"快递包裹"},
    u"山西": {'ticket_initial': 9, 'package_type': u"快递包裹"},
    u"陕西": {'ticket_initial': 9, 'package_type': u"快递包裹"},
    u"贵州": {'ticket_initial': 9, 'package_type': u"快递包裹"},
    u"云南": {'ticket_initial': 9, 'package_type': u"快递包裹"},
    u"重庆": {'ticket_initial': 9, 'package_type': u"快递包裹"},
    u"四川": {'ticket_initial': 9, 'package_type': u"快递包裹"},
    u"北京": {'ticket_initial': 9, 'package_type': u"快递包裹"},
    u"天津": {'ticket_initial': 9, 'package_type': u"快递包裹"},
    u"辽宁": {'ticket_initial': 9, 'package_type': u"快递包裹"},
    u"黑龙江": {'ticket_initial': 9, 'package_type': u"快递包裹"},

    u"广东": {'ticket_initial': 1, 'package_type': u"标准快递"},

    u"内蒙古": {'ticket_initial': 1, 'package_type': u"标准快递"},
    u"甘肃": {'ticket_initial': 1, 'package_type': u"标准快递"},
    u"青海": {'ticket_initial': 1, 'package_type': u"标准快递"},
    u"宁夏": {'ticket_initial': 1, 'package_type': u"标准快递"},
    u"吉林": {'ticket_initial': 1, 'package_type': u"标准快递"},
    u"西藏": {'ticket_initial': 1, 'package_type': u"标准快递"},
    u"新疆": {'ticket_initial': 1, 'package_type': u"标准快递"},
}

PROVINCE_NAMES = [City.normalize_province(name) for name in PROVINCE_INFO_MAP.keys()]

ADDRESS_LIMIT_FIXTURES = [u'维吾尔自治区', u'壮族自治区', u'回族自治区', u'古自治区', u'自治区', u'特别行政区', u'省', u'市',
                          u'盟', u'县', u'区']

ITEM_NAME_RE = re.compile(
    ur"^.*?((([123一二三])|([4四]))段|(\d+)g)$", flags=re.U | re.I)

ITEM_NAME_MAP_INFO = {
    u"爱他美1段": {
        "net_weight": 0.9,
        "gross_weights": {
            4: 4.35,
            6: 7,
        },
        "price_per_kg": 89.06,
        "full_name": u"爱他美奶粉1段900g",
    },
    u"爱他美2段": {
        "net_weight": 0.9,
        "gross_weights": {
            4: 4.35,
            6: 7,
        },
        "price_per_kg": 89.06,
        "full_name": u"爱他美奶粉2段900g",
    },
    u"爱他美3段": {
        "net_weight": 0.9,
        "gross_weights": {
            4: 4.35,
            6: 7,
        },
        "price_per_kg": 89.06,
        "full_name": u"爱他美奶粉3段900g",
    },
    u"爱他美4段": {
        "net_weight": 0.8,
        "gross_weights": {
            4: 3.90,
            6: 6,
        },
        "price_per_kg": 86.50,
        "full_name": u"爱他美奶粉4段800g",
    },
    u"牛栏1段": {
        "net_weight": 0.9,
        "gross_weights": {
            4: 4.35,
            6: 7,
        },
        "price_per_kg": 89.06,
        "full_name": u"牛栏奶粉1段900g",
    },
    u"牛栏2段": {
        "net_weight": 0.9,
        "gross_weights": {
            4: 4.35,
            6: 7,
        },
        "price_per_kg": 86.50,
        "full_name": u"牛栏奶粉2段900g",
    },
    u"牛栏3段": {
        "net_weight": 0.9,
        "gross_weights": {
            4: 4.35,
            6: 7,
        },
        "price_per_kg": 86.50,
        "full_name": u"牛栏奶粉3段900g",
    },
    u"牛栏4段": {
        "net_weight": 0.8,
        "gross_weights": {
            4: 3.90,
            6: 6,
        },
        "price_per_kg": 94.30,
        "full_name": u"牛栏奶粉4段800g",
    },
}


def calculate_item_info(n_row, item_name, item_count):
    item_name = "".join(item_name.strip().split()).decode("utf8")
    if not item_name in ITEM_NAME_MAP_INFO:
        raise Exception, u"第%d行包含未注册商品名称: %s" % (n_row + 1, item_name)
    info = ITEM_NAME_MAP_INFO[item_name]
    if not item_count in info["gross_weights"]:
        raise Exception, u"第%d行商品[%s]包含未注册数量:%d" % (
            n_row + 1, item_name, item_count)
    return info["net_weight"] * item_count * info["price_per_kg"], info["net_weight"] * item_count, \
        info["gross_weights"][item_count], info["price_per_kg"], \
        info["full_name"] if "full_name" in info else item_name


def calculate_item_info_from_db(n_row, item_name, item_count):
    item_name = "".join(item_name.strip().split()).decode("utf8")
    search_result = ProductInfo.find_product_and_weight(item_name, item_count)
    if not search_result:
        raise Exception, u"第%d行包含未注册商品名称和箱件数 %s[%d件]" % (
            n_row + 1, item_name, item_count)
    product_info, gross_weight_per_box = search_result
    return product_info.net_weight * item_count * product_info.price_per_kg, product_info.net_weight * item_count, \
        gross_weight_per_box, product_info.price_per_kg, \
        product_info.full_name if product_info.full_name else item_name


def calculate_item_info_from_db_without_product_info(n_row, item_name, item_count):
    item_name = "".join(item_name.strip().split()).decode("utf8")
    product_info = ProductInfo.query.filter(and_(ProductInfo.name==item_name, ProductInfo.deprecated==False)).first()
    if not product_info:
        raise Exception, u"第%d行包含未注册商品: %s" % (n_row + 1, item_name)

    # from sqlalchemy import inspect
    # for c in inspect(product_info).mapper.column_attrs:
    #     print c.key, getattr(product_info, c.key)
    if product_info.unit_price is None or product_info.gross_weight is None or product_info.unit_per_item is None \
            or product_info.tax_code is None or product_info.billing_unit is None \
            or product_info.billing_unit_code is None or product_info.specification is None \
            or not product_info.full_name:
        raise Exception, u"第%d行商品 [%s] 注册信息不完整" % (n_row + 1, item_name)

    waybill_name = product_info.waybill_name if product_info.waybill_name else product_info.full_name

    return product_info.unit_price * product_info.unit_per_item * item_count, product_info.net_weight * item_count, \
        product_info.gross_weight * item_count, product_info.unit_price, \
        product_info.full_name, \
        product_info.net_weight, product_info.tax_code, product_info.billing_unit, product_info.billing_unit_code, \
        product_info.unit_per_item, product_info.specification, waybill_name


class NoTextImageWriter(ImageWriter):

    def __init__(self):
        super(NoTextImageWriter, self).__init__()

    def _paint_text(self, xpos, ypos):
        pass

def random_date():
    now = datetime.datetime.now()
    start = now + datetime.timedelta(days=-21)
    end = now + datetime.timedelta(days=-1)
    delta = end - start
    int_delta = (delta.days * 24 * 60 * 60) + delta.seconds
    random_second = random.randrange(int_delta)
    candidate = start + datetime.timedelta(seconds=random_second)
    if 8 <= candidate.hour <= 18:
        pass
    else:
        candidate = candidate.replace(hour=random.randrange(8, 18))
    return candidate

TICKET_SELECTIONS = [
    {
        'template': 'tickets/tesco.html',
        'images': ['static/img/tickets/TESCO-1.jpg', 'static/img/tickets/TESCO-2.jpg'],
        'serial_number_len': 13
    },
    {
        'template': 'tickets/asda.html',
        'images': ['static/img/tickets/ASDA-1.jpg',],
        'serial_number_len': 13
    },
    {
        'template': 'tickets/morrisons.html',
        'images': ['static/img/tickets/Morrisons-1.jpg', 'static/img/tickets/Morrisons-2.jpg', ],
        'serial_number_len': 13
    },
    {
        'template': 'tickets/quality_safe.html',
        'images': ['static/img/tickets/QualitySafe-1.jpg', 'static/img/tickets/QualitySafe-2.jpg', ],
        'serial_number_len': 8
    },
]

def generate_tickets_from_mapping_file(input_xlsx, mapping_xlsx, output_dir):

    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    if os.path.isdir(input_xlsx):
        input_xlsxs = map(lambda x:os.path.join(input_xlsx, x),
                          filter(lambda x:x.lower().endswith('.xlsx'), os.listdir(input_xlsx)))

    else:
        input_xlsxs = [input_xlsx]
    input_dfs = [pd.read_excel(input_xlsx, skiprows=[0, ], converters={
        u'分运单号': lambda x: str(x),
        u'货物名称': lambda x: str(x),
    }) for input_xlsx in input_xlsxs]

    input_df = pd.concat(input_dfs)

    print >> sys.stderr, "%d total tracking numbers" % len(input_df[u'分运单号'].drop_duplicates().index)

    mapping_df = pd.read_excel(mapping_xlsx, converters={
        u'货物名称': lambda x: str(x),
        u'小票名称': lambda x: str(x),
        u'小票价格': lambda x: float(x),
    })
    duplicates = mapping_df.groupby(u'货物名称').filter(lambda x: len(x) > 1).drop_duplicates(subset=u'货物名称')
    if len(duplicates) > 0:
        print >> sys.stderr, duplicates
        raise Exception, "duplicates"

    combined_df = pd.merge(input_df, mapping_df, on=u'货物名称', how='left')
    unregisted_products = set()
    for column in (u"小票名称", u"小票价格"):
        null_valued = pd.isnull(combined_df[column])
        if null_valued.any():
            product_name_null_valued = combined_df[null_valued][u'货物名称'].drop_duplicates() \
                .map(lambda x: str(x)).tolist()
            unregisted_products |= set(product_name_null_valued)
    if len(unregisted_products) > 0:
        with open(os.path.join(output_dir, 'products.txt'), 'w') as f:
            for product in sorted(unregisted_products):
                print >> f, product
        raise Exception, "contains unregisted product"

    for input_xlsx, input_df in zip(input_xlsxs, input_dfs):
        print >> sys.stderr, "processing: ", input_xlsx
        print >> sys.stderr, "%d tracking numbers" % len(input_df[u'分运单号'].drop_duplicates().index)
        combined_df = pd.merge(input_df, mapping_df, on=u'货物名称', how='left')

        ticket_info = {
            'groups': combined_df[[u'分运单号', u"小票名称", u'件数', u"小票价格"]].groupby(u'分运单号'),
            'item_column': u"小票名称",
            'count_column': u'件数',
            'price_column': u"小票价格",
        }

        ticket_dir = os.path.join(output_dir, os.path.splitext(os.path.basename(input_xlsx))[0])
        if not os.path.exists(ticket_dir):
            os.makedirs(ticket_dir)

        generate_tickets(ticket_info, ticket_dir, suffix='.jpg')

        if os.path.exists(ticket_dir):
            shutil.make_archive(ticket_dir, 'zip', ticket_dir)
        shutil.rmtree(ticket_dir)

def generate_tickets(ticket_info, ticket_dir, suffix='.x.jpg'):
    item_column = ticket_info['item_column']
    count_column = ticket_info['count_column']
    price_column = ticket_info['price_column']
    faker = Faker()

    def format_number(x):
        if isinstance(x, float) or isinstance(x, int):
            return "%.2f" % x
        return map(lambda x:"%.2f" % x, x)

    def random_context(selections):
        supermarket = random.choice(selections)
        supermarket['image'] = random.choice(supermarket['images'])
        return supermarket

    generated_count = 0
    for name, group in ticket_info['groups']:
        filename = os.path.join(ticket_dir, name + suffix)
        env = Environment(loader=FileSystemLoader('templates'))

        total = 0
        total_items = 0
        for price, count in zip(group[price_column], group[count_column]):
            total += price * count
            total_items += count

        paid = int(math.ceil(total / 10.0)) * 10
        change = paid - total

        context = random_context(TICKET_SELECTIONS)
        context.update({
            'breakdown': zip(group[item_column], group[count_column], format_number(group[price_column])),
            'total': format_number(total),
            'paid': format_number(paid),
            'change': format_number(change),
            'points': int(total * 100),
            'timestamp': random_date(),
            'serial_number': faker.ean(length=context['serial_number_len']),
            'total_items': total_items,
        })
        template = env.get_template(context['template'])

        output_from_parsed_template = template.render(context)

        png_data = HTML(string=output_from_parsed_template, base_url='.').write_png(
            stylesheets=["static/css/ticket_style.css"], resolution=150)

        im = Image(blob=png_data)
        im.trim(Color('white'))
        im.format = 'jpeg'
        im.trim(Color('black'))
        im.save(filename=filename)
        generated_count += 1

    print >> sys.stderr, "%d tickets generated" % generated_count


def generate_pdf(ticket_number, filename, context, tmpdir):
    if not os.path.exists(tmpdir):
        os.makedirs(tmpdir)

    bot_image = os.path.join(tmpdir, 'bot_barcode_trim.png')
    top_image = os.path.join(tmpdir, 'top_barcode_trim.png')
    if os.path.exists(bot_image):
        os.remove(bot_image)
    if os.path.exists(top_image):
        os.remove(top_image)

    Code128(ticket_number, writer=NoTextImageWriter()).save(os.path.join(tmpdir, 'top_barcode'), options={
        'module_height': 5,
        'text_distance': 0.5,
        'quiet_zone': 1,
        'dpi': 1200,
        'font_size': 20,
    })

    im = Image(filename=os.path.join(tmpdir, 'top_barcode.png'))
    im.trim()
    im.save(filename=top_image)

    Code128(ticket_number, writer=NoTextImageWriter()).save(os.path.join(tmpdir, 'bot_barcode'), options={
        'module_height': 5,
        'text_distance': 0.5,
        'quiet_zone': 1,
        'dpi': 1200,
        'font_size': 20,
    })

    im = Image(filename=os.path.join(tmpdir, 'bot_barcode.png'))
    im.trim()
    im.save(filename=bot_image)

    if not os.path.exists(top_image) or not os.path.exists(bot_image):
        raise Exception, "Image failed to create"

    env = Environment(loader=FileSystemLoader('templates'))
    template = env.get_template('barcode.html')
    context['time'] = datetime.datetime.now()
    context['bot_image'] = bot_image
    context['top_image'] = top_image
    output_from_parsed_template = template.render(context)
    #
    # with codecs.open(filename + ".html", "wb", encoding='utf8') as fh:
    #     fh.write(output_from_parsed_template)

    HTML(string=output_from_parsed_template, base_url='.').write_pdf(
        filename, stylesheets=["static/css/style.css"])


def fetch_ticket_number(n_row, receiver_city):
    city_name = "".join(receiver_city.strip().split())
    cities = City.find_province_path(city_name)
    if not cities:
        raise Exception, "Cannot find province: %s at row %d" % (
            city_name, n_row)
    if not cities[0].name in PROVINCE_INFO_MAP:
        raise Exception, "Post to province %s is not supported: %s at row %d" % (
            city_name, n_row)
    info = PROVINCE_INFO_MAP[cities[0].name]
    order = Order.pick_unused()
    if order is None:
        raise Exception, u"订单号不足"
    province_name, municipal_name, address_header = City.normalize_province_path(
        cities)
    return info['package_type'], order, province_name, municipal_name, address_header


def process_row(n_row, in_row, barcode_dir, tmpdir, job=None, ticket_number_generator=None):
    p_data = []
    sender_name = in_row[u'发件人名字']
    sender_phone = in_row[u'发件人电话号码']
    sender_address = in_row[u'发件人地址']
    receiver_name = in_row[u'收件人名字（中文）']
    receiver_mobile = in_row[u'收件人手机号（11位数）']
    receiver_address = in_row[u'收件人地址（无需包括省份和城市）']
    receiver_city = in_row[u'收件人城市（中文）']
    receiver_post_code = in_row[u'收件人邮编']
    n_package = in_row[u'包裹数量']
    package_weight = in_row[u'包裹重量（公斤）']
    length = in_row[u'长（厘米）']
    width = in_row[u'宽（厘米）']
    height = in_row[u'高（厘米）']
    id_number = in_row[u'身份证号(EMS需要)']

    for check_field in (sender_name, sender_phone, sender_address, receiver_name, receiver_mobile, receiver_address,
                        receiver_city, receiver_post_code, id_number):
        if pd.isnull(check_field) or not isinstance(check_field, basestring) or not check_field.strip():
            raise Exception, u"第%d行数据不完整,请更正" % n_row

    if pd.isnull(n_package) or not isinstance(n_package, int) or n_package < 1:
        raise Exception, u"第%d行包裹数量异常" % n_row

    sender_name = "".join(sender_name.split())
    sender_address = "".join(sender_address.split())
    sender_phone = "".join(sender_phone.split())
    receiver_name = "".join(receiver_name.split())
    receiver_mobile = "".join(receiver_mobile.split())
    receiver_address = "".join(receiver_address.split())
    receiver_city = "".join(receiver_city.split())
    receiver_post_code = "".join(receiver_post_code.split())
    id_number = "".join(id_number.split())

    package_type, order, receiver_province, receiver_municipal, receiver_address_header = \
        fetch_ticket_number(n_row, receiver_city)
    receiver_city = receiver_municipal
    receiver_address = receiver_address_header + receiver_address

    pc_text = receiver_province + receiver_municipal
    receiver_province_city_font_size = "3" if len(
        pc_text) <= 10 else "2.5" if len(pc_text) <= 15 else "2"

    if not ticket_number_generator:
        order.used = True
        order.used_time = datetime.datetime.utcnow()
        order.sender_address = ", ".join(
            (sender_name, sender_address, sender_phone))
        order.receiver_address = ", ".join(
            (receiver_address, receiver_city, receiver_post_code))
        order.receiver_mobile = receiver_mobile
        order.receiver_id_number = id_number
        order.receiver_name = receiver_name
        if job:
            order.job = job
            job.version = "v3"
        ticket_number = order.order_number
    else:
        ticket_number = ticket_number_generator.next()
    full_address = "".join(filter(
        lambda x: x.strip(), (receiver_province, receiver_city, receiver_address)))

    p_data_list = []

    item_names = []
    total_price = 0
    total_item_count = 0
    total_net_weight = 0
    total_gross_weight = 0
    for i in xrange(n_package):
        suffix = "" if i == 0 else ".%d" % i
        item_name = in_row[u'申报物品%d(英文）' % (i + 1)]
        item_count = in_row[u'数量%s' % suffix]
        unit_price = in_row[u'物品单价（英镑）%s' % suffix]

        if item_name is None or pd.isnull(item_name):
            raise Exception, u"第%d行第%d个商品名称为空" % (n_row, i + 1)
        item_name = str(item_name).strip()

        sub_total_price, net_weight, gross_weight, unit_price, item_full_name, net_weight_per_item, tax_code, \
        billing_unit, billing_unit_code, unit_per_item, specification, waybill_name \
            = calculate_item_info_from_db_without_product_info(n_row, item_name, item_count)

        item_names.append(u"%s\u2736%d" % (waybill_name, item_count))
        total_price += sub_total_price
        total_item_count += item_count
        total_net_weight += net_weight
        total_gross_weight += gross_weight

        p_data_list.append([
            ticket_number, sender_name, sender_address, sender_phone, receiver_name, receiver_mobile, receiver_city if receiver_city else receiver_province,
            receiver_post_code, full_address, item_full_name, item_count, sub_total_price, gross_weight, item_full_name,
            net_weight, unit_price, u"CNY", id_number,
            City.denormalize_province(receiver_province),
            City.denormalize_municipality(receiver_city if receiver_city else receiver_province)
        ])


    # for p in p_data_list:
    #     p[10] = total_item_count
    #     p[11] = total_price
    #     p[12] = total_gross_weight
    #     p_data.append(p)
    p_data = p_data_list

    total_price = "%.2f" % total_price
    if total_price.endswith(".00") and len(total_price) > 3:
        total_price = total_price[:-3]

    item_names = ", ".join(item_names)

    generate_pdf(ticket_number, os.path.join(
        barcode_dir, '%s.pdf' % ticket_number), locals(), tmpdir)

    return ticket_number, pd.DataFrame(p_data, columns=[
        u'快件单号', u'发件人', u'发件人地址', u'电话号码', u'收件人', u'电话号码.1', u'城市',
        u'邮编', u'收件人地址', u'内件名称', u'数量', u'总价（元）', u'毛重（KG）', u'物品名称',
        u'数量.1', u'单价', u'币别', u'备注', 'province', 'city'
    ])


def normalize_columns(in_df):
    in_df.columns = map(lambda x: "".join(x.strip().split()), in_df.columns)


def xls_to_orders(input, output, tmpdir, percent_callback=None, job=None, test_mode=False):
    if percent_callback:
        percent_callback(0)
    in_df = pd.read_excel(input, converters={
        u'发件人电话号码': lambda x: str(x),
        u'收件人邮编': lambda x: str(x),
        u'收件人手机号\n（11位数）': lambda x: str(x),
        u'身份证号\n(EMS需要)': lambda x: str(x),
        u'收件人手机号（11位数）': lambda x: str(x),
        u'身份证号(EMS需要)': lambda x: str(x),
        u'包裹数量': lambda x: int(x),
    })
    if 'MAX_ORDER_PER_BATCH' in current_app.config \
            and len(in_df.index) > current_app.config['MAX_ORDER_PER_BATCH']:
        raise Exception, u"该批次个数(%d)超过最大订单数: %d" % \
                         (len(in_df.index), current_app.config['MAX_ORDER_PER_BATCH'])
    normalize_columns(in_df)

    package_columns = [u"报关单号", u'总运单号', u'袋号', u'快件单号', u'发件人', u'发件人地址',
                       u'电话号码', u'收件人', u'电话号码.1', u'城市', u'邮编', u'收件人地址', u'内件名称',
                       u'数量', u'总价（元）', u'毛重（KG）', u'税号', u'物品名称', u'品牌', u'数量.1',
                       u'单位', u'单价', u'币别', u'备注', 'province', 'city']

    package_df = pd.DataFrame([], columns=package_columns)
    package_data = [package_df]

    barcode_dir = os.path.join(output, "barcode")
    if not os.path.exists(barcode_dir):
        os.makedirs(barcode_dir)

    ticket_numbers = []
    ticket_number_set = set()
    test_ticket_number_generator = None
    if test_mode:
        def ticket_number_generator():
            start_number = 1
            while True:
                yield "TEST%s" % str(start_number).zfill(8)
                start_number += 1

        test_ticket_number_generator = ticket_number_generator()
        if job:
            job.version = "test_mode"
    for index, in_row in in_df.iterrows():
        ticket_number, p_data = process_row(
            index, in_row, barcode_dir, tmpdir, job, test_ticket_number_generator)
        if ticket_number in ticket_number_set:
            raise Exception, u"同批次单号%s重复，请联系客服！" % ticket_number
        ticket_number_set.add(ticket_number)
        ticket_numbers.append(ticket_number)
        package_data.append(p_data)
        if percent_callback:
            percent_callback(int(index * 100.0 / len(in_df.index)))

    waybills = []
    total_page_number = 0
    merger = PdfFileMerger()
    for ticket_number in ticket_numbers:
        pdf_file = os.path.join(barcode_dir, "%s.pdf" % ticket_number)
        if not os.path.exists(pdf_file):
            raise Exception, "Failed to generate pdf: %s" % ticket_number
        pdf_file_reader = PdfFileReader(file(pdf_file, 'rb'))
        page_number = pdf_file_reader.getNumPages()
        waybills.append({
            'tracking_no': ticket_number,
            'start_page': total_page_number,
            'end_page' : total_page_number + page_number,
        })
        total_page_number += page_number
        merger.append(pdf_file_reader)
    merger.write(os.path.join(output, u"面单.pdf".encode('utf8')))
    with open(os.path.join(output, "waybills.json"), 'w') as outfile:
        json.dump(waybills, outfile)
    shutil.rmtree(barcode_dir)

    package_final_df = pd.concat(package_data, ignore_index=True)
    package_final_df[u'税号'] = '01010700'
    package_final_df[u'单位'] = u'千克'
    package_final_df.index += 1
    package_final_df.to_excel(os.path.join(output, u"机场报关单.xlsx".encode('utf8')),
                              columns=package_columns, index_label="NO")


    if percent_callback:
        percent_callback(100)


def read_order_numbers(inxlsx):
    columns = [u'提取单号', u'分运单号', u'快件单号', u'物流运单编号']
    df = pd.read_excel(inxlsx, converters={
        key: lambda x: str(x) for key in columns
    })
    column = None
    for key in columns:
        if key in df:
            column = key
    if not column:
        raise Exception, u"输入Excel格式错误"
    order_numbers = df[column].unique()
    if len(order_numbers) <= 0:
        raise Exception, u"输入[%s]列为空" % column
    return order_numbers


def generate_customs_df(route_config, version, package_df):
    route_name = route_config['name']
    if version <> "v3":
        raise Exception, "Version not supported for generate_customs_df: %s" % version

    package_df["Sequence"] = range(1, len(package_df.index) + 1)

    customs_columns = [u'序号', u'分运单号', u'货物品名', u'件数', u'提单重量', u'派送重量',
                       u'数量', u'实际数量', u'单位',
                       u'货币编码', u'单价', u'个人完税税号', u'型号', u'国别代码', u'原产国',
                       u'HS编码', u'收件人ID', u'收件人', u'地址', u'收件人电话', u'TO',
                       u'落地配单号', u'寄件人公司', u'寄件人', u'寄件人地址', u'寄件人电话',
                       u'FROM', u'货主城市']
    customs_df = pd.DataFrame([], columns=customs_columns)
    for column, p_column in ((u'分运单号', u'快件单号'),
                             (u'货物品名', u'内件名称'),
                             (u'派送重量', u'毛重（KG）'),
                             (u'数量', u'数量'),
                             (u'实际数量', u'数量'),
                             (u'收件人ID', u'备注'),
                             (u'收件人', u'收件人'),
                             (u'地址', u'收件人地址'),
                             (u'收件人电话', u'电话号码.1'),
                             (u'TO', u'province'),
                             (u'落地配单号', u'快件单号'),
                             (u'寄件人', u'发件人'),
                             (u'寄件人地址', u'发件人地址'),
                             (u'寄件人电话', u'电话号码'),
                             (u'货主城市', u'province'),
                             ("Sequence", "Sequence")):
        customs_df[column] = package_df[p_column]

    #fill in bc product info
    product_info_df = pd.read_sql_query(ProductInfo.query.filter(ProductInfo.full_name.in_(
        tuple(set(customs_df[u'货物品名'].map(lambda x: str(x)).tolist())))).statement, db.session.bind)
    columns_to_delete = product_info_df.columns
    product_info_df.rename(columns={'full_name': u'货物品名'}, inplace=True)
    customs_df = pd.merge(customs_df, product_info_df, on=u'货物品名')
    product_info_columns = [(u"单位", "billing_unit"),
                            (u"单价", "unit_price"),
                            (u"个人完税税号", "tax_code"),
                            (u"型号", "specification")]
    # check if any is empty
    for column, _column in product_info_columns \
            + [(u"单个物品申报数量", "unit_per_item"),
               (u"小票名称", "ticket_name"),
               (u"小票价格", "ticket_price")]:
        null_valued = pd.isnull(customs_df[_column])
        if null_valued.any():
            product_name_null_valued = customs_df[null_valued][u'货物品名'].drop_duplicates() \
                .map(lambda x: str(x)).tolist()
            raise Exception, u"如下商品的注册信息未包含必须字段[%s]: %s" % \
                             (column, ", ".join(product_name_null_valued))

    ticket_info = {
        'groups': customs_df[[u'分运单号', "ticket_name", u'实际数量', "ticket_price"]].groupby(u'分运单号'),
        'item_column': 'ticket_name',
        'count_column': u'实际数量',
        'price_column': 'ticket_price',
    }

    for column, p_column in product_info_columns:
        customs_df[column] = customs_df[p_column]

    def customs_column_filter(row):
        name = row[u"货物品名"] if pd.isnull(row["report_name"]) else row["report_name"]
        row[u"货物品名"] = "%s*%d" % (name, row[u"实际数量"])
        row[u"数量"] = row[u"实际数量"] * row["unit_per_item"]
        return row

    customs_df = customs_df.apply(customs_column_filter, axis=1)

    for column in columns_to_delete:
        if column in customs_df:
            del customs_df[column]

    customs_df.sort_values(by=["Sequence"], inplace=True)

    #index create
    index_df = customs_df[[u"分运单号"]].copy()
    index_df.drop_duplicates(inplace=True)
    index_df["order"] = range(1, len(index_df.index) + 1)
    customs_df = pd.merge(customs_df, index_df, on=u'分运单号')
    customs_df[u"序号"] = customs_df['order']
    del customs_df["order"]

    #fixed items
    customs_df[u"件数"] = "1"
    customs_df[u"货币编码"] = "RMB"
    customs_df[u"国别代码"] = "303"
    customs_df[u"原产国"] = u"英国"
    customs_df[u"FROM"] = "Man"
    customs_df[u'HS编码'] = "1901101000"

    #sort
    customs_df.sort_values(by=[u"序号", u'分运单号'], inplace=True)

    del customs_df["Sequence"]
    del package_df["Sequence"]

    return customs_df, ticket_info

def map_full_name_to_report_name(data_df, column_name):
    if not column_name in data_df.columns:
        raise Exception, "%s not in header" % column_name
    product_info_df = pd.read_sql_query(ProductInfo.query.filter(ProductInfo.full_name.in_(
        tuple(set(data_df[column_name].map(lambda x: str(x)).tolist())))).statement, db.session.bind)
    columns_to_delete = product_info_df.columns
    product_info_df.rename(columns={'full_name': column_name}, inplace=True)
    data_df = pd.merge(data_df, product_info_df, on=column_name)
    data_df[column_name] = data_df.apply(lambda row:row['report_name'] if row['report_name'] else row[column_name],
                                         axis=1)
    for column in columns_to_delete:
        if column in data_df:
            del data_df[column]
    return data_df

def remap_customs_df(customs_final_df):
    wb = load_workbook(os.path.join(os.path.dirname(os.path.realpath(__file__)), 'cc_header.xlsx'))
    ws = wb["Sheet1"]
    row_count = 0
    for r in dataframe_to_rows(customs_final_df, index=False, header=False):
        row_count += 1
        ws.append(r)

    # merge cell for this one
    # base_index = 7
    # last_value = 0
    # last_row_num = None
    # columns = (1, 2, 4, 15, 16, 17, 18, 19, 20, 22, 23, 24, 26)
    #for row_num in range(base_index, base_index + row_count):
    #    rd = ws.row_dimensions[row_num]
    #    rd.height = 50

        # is_last_row = (row_num == base_index + row_count - 1)
        #
        # package_index = int(ws.cell(row=row_num, column=1).value)
        # assert (package_index > 0)
        # if last_value <= 0:
        #     last_value = package_index
        #     last_row_num = row_num
        # else:
        #     if is_last_row or last_value != package_index:
        #         if row_num > last_row_num + 1 or (is_last_row and row_num > last_row_num and last_value == package_index):
        #             start_row = last_row_num
        #             end_row = row_num if is_last_row else row_num - 1
        #             for _row_num in range(start_row, end_row):
        #                 for column in columns:
        #                     first_value = ws.cell(row=_row_num, column=column).value
        #                     second_value = ws.cell(row=end_row, column=column).value
        #                     assert ((isinstance(first_value, float) and isinstance(second_value, float) and
        #                              math.isnan(first_value) and math.isnan(second_value))
        #                             or (first_value == second_value))
        #             for column in columns:
        #                 ws.merge_cells(start_row=start_row, start_column=column,
        #                                end_row=end_row, end_column=column)
        #         last_value = package_index
        #         last_row_num = row_num
    return wb

def retract_from_order_numbers(download_folder, order_numbers, output, route_config, retraction=None):
    route_name = route_config['name']
    route_code = route_config['code']

    if not output is None:
        waybill_dir = os.path.join(output, u"面单")
        ticket_dir = os.path.join(output, u"小票")
        if not os.path.exists(waybill_dir):
            os.makedirs(waybill_dir)
        if not os.path.exists(ticket_dir):
            os.makedirs(ticket_dir)

    # find all jobs and job to order number map
    receiver_sig_to_order_numbers = {}
    all_order_numbers = set()
    uuid_to_order_numbers = {}
    job_versions = {}
    for i, order_number in enumerate(order_numbers):
        order_number = str(order_number).strip()
        if order_number in all_order_numbers:
            continue
        else:
            all_order_numbers.add(order_number)
        order = Order.find_by_order_number(order_number)
        if not order:
            raise Exception, u"第%d行包含未上载订单号: %s" % (i + 1, order_number)
        if not order.used:
            raise Exception, u"第%d行包含未使用订单号: %s" % (i + 1, order_number)
        if retraction is not None and order.retraction is not None:
            raise Exception, u"第%d行订单号已被提取: %s, 提取信息为: Uuid [%s], 时间 [%s]" % \
                             (i + 1, order_number, order.retraction.uuid,
                              time_to_filename(order.retraction.timestamp))
        receiver_sig = order.receiver_id_number
        if not receiver_sig in receiver_sig_to_order_numbers:
            receiver_sig_to_order_numbers[receiver_sig] = []
        if len(receiver_sig_to_order_numbers[receiver_sig]) >= route_config['max_order_number_per_receiver']:
            raise Exception, u"单个收件人超过最大订单数(%d): 第%d行订单(%s)与[ %s ]包含相同证件号码(%s), 收件人: %s" % \
                             (route_config['max_order_number_per_receiver'], i + 1, order_number,
                              " / ".join(["第%d行订单(%s)" % (x + 1, y)
                                          for x, y in receiver_sig_to_order_numbers[receiver_sig]]),
                              receiver_sig, order.receiver_name)
        receiver_sig_to_order_numbers[receiver_sig].append((i, order_number))
        uuid = str(order.job.uuid)
        if not uuid in uuid_to_order_numbers:
            uuid_to_order_numbers[uuid] = set()
        uuid_to_order_numbers[uuid].add(order_number)
        job_versions[uuid] = order.job.version if order.job.version else "v1"
        if retraction:
            order.retraction = retraction

    version_to_dfs = {}

    for uuid, order_number_set in uuid_to_order_numbers.items():
        version = job_versions[uuid]
        if not version in version_to_dfs:
            version_to_dfs[version] = {'package_dfs': [], 'customs_dfs' : []}
        package_dfs = version_to_dfs[version]['package_dfs']
        job_file = os.path.join(download_folder, uuid, uuid + '.zip')
        if not os.path.exists(job_file):
            raise Exception, u"历史数据丢失:%s" % uuid
        with zipfile.ZipFile(job_file) as z:
            if version == "v3":
                package_df = pd.read_excel(z.open(u"机场报关单.xlsx"), index_col=0, converters={
                    u'快件单号': lambda x: str(x),
                    u'电话号码': lambda x: str(x),
                    u'电话号码.1': lambda x: str(x),
                    u'邮编': lambda x: str(x),
                    u'税号': lambda x: str(x),
                    u'备注': lambda x: str(x),
                })

                sub_package_df = package_df[
                    package_df[u"快件单号"].isin(order_number_set)]

                #output waybill
                if output:
                    waybills = json.load(z.open('waybills.json'))
                    pdf_data = StringIO(z.open(u"面单.pdf").read())
                    pdf_data.seek(0)
                    pdf = PdfFileReader(pdf_data)
                    page_count = pdf.getNumPages()
                    for waybill in waybills:
                        if waybill['tracking_no'] in order_number_set:
                            if waybill['end_page'] > page_count or waybill['start_page'] >= page_count:
                                raise Exception, "Waybill page length %d-%d larger than pdf length %d" % \
                                                 (waybill['start_page'], waybill['end_page'], page_count)
                            out_pdf = PdfFileWriter()
                            for i in xrange(waybill['start_page'], waybill['end_page']):
                                out_pdf.addPage(pdf.getPage(i))

                            pdf_content = StringIO()
                            out_pdf.write(pdf_content)
                            pdf_content.seek(0)
                            images = convert_from_bytes(pdf_content.read(), dpi=50)
                            if images:
                                images[0].save(os.path.join(waybill_dir, waybill['tracking_no'] + '.y.jpg'))
                            else:
                                raise Exception, "No jpg waybill generated for %s" % waybill['tracking_no']
            else:
                raise Exception, "Version not supported %s" % version

            package_dfs.append(sub_package_df)

    for version, data in version_to_dfs.iteritems():
        package_dfs = data['package_dfs']

        def validate_route(package_df):
            if version == "v3":
                product_col = u"内件名称"
                order_number_col = u'快件单号'
                count_col = u'数量'
            else:
                raise Exception, "Version not supported: %s" % version

            products_exclude = route_config['products_exclude'] if 'products_exclude' in route_config else []
            for product_exclude in products_exclude:
                product_exclude = product_exclude.strip()
                if product_exclude:
                    excluded = package_df[product_col].str.contains(product_exclude)
                    if excluded.any():
                        order_numbers_excluded = package_df[excluded][order_number_col].map(lambda x:str(x)).tolist()
                        raise Exception, u"如下订单号包含违禁产品[%s]: %s" % \
                                         (product_exclude, ", ".join(order_numbers_excluded))

        package_final_df = pd.concat(package_dfs, ignore_index=True)
        package_final_df.index += 1

        validate_route(package_final_df)

        if output:
            if version == "v3":
                customs_final_df, ticket_info = generate_customs_df(route_config, version, package_final_df)

                generate_tickets(ticket_info, ticket_dir)

                wb = remap_customs_df(customs_final_df)
                wb.save(os.path.join(output, u"晋江申报单.xlsx".encode('utf8')))

                del package_final_df["province"]
                del package_final_df["city"]
                package_final_df.to_excel(os.path.join(
                    output, u"机场报关单.xlsx".encode('utf8')), index_label="NO")

                if os.path.exists(waybill_dir):
                    shutil.make_archive(waybill_dir, 'zip', waybill_dir)
                shutil.rmtree(waybill_dir)
                if os.path.exists(ticket_dir):
                    shutil.make_archive(ticket_dir, 'zip', ticket_dir)
                shutil.rmtree(ticket_dir)

            else:
                raise Exception, "Version not supported too %s" % version

    return package_final_df

def load_order_info(download_folder, order, route_config):
    version = order.job.version if order.job.version else "v1"
    if version <> "v3":
        raise Exception, "Error: Version"
    route_code = route_config['code']
    product_col = u"内件名称"
    uuid = str(order.job.uuid)
    job_file = os.path.join(download_folder, uuid, uuid + '.zip')
    if not os.path.exists(job_file):
        raise Exception, "Error: Missing File"
    with zipfile.ZipFile(job_file) as z:
        package_df = pd.read_excel(z.open(u"机场报关单.xlsx"), index_col=0, converters={
            u'快件单号': lambda x: str(x),
            u'电话号码': lambda x: str(x),
            u'电话号码.1': lambda x: str(x),
            u'邮编': lambda x: str(x),
            u'税号': lambda x: str(x),
            u'备注': lambda x: str(x),
        })
        sub_package_df = package_df[package_df[u"快件单号"] == order.order_number]
        if len(sub_package_df.index) <= 0:
            raise Exception, "Error: Empty Record"

        products_exclude = route_config['products_exclude'] if 'products_exclude' in route_config else []
        if products_exclude:
            for product_exclude in products_exclude:
                product_exclude = product_exclude.strip()
                if product_exclude:
                    excluded = sub_package_df[product_col].str.contains(product_exclude)
                    if excluded.any():
                        raise Exception, "Error: Product Excluded"
        pieces = sub_package_df[u"数量"].sum()

    return sub_package_df, pieces

if __name__ == "__main__":
    parser = OptionParser()
    parser.add_option("-i", "--input", dest="input",
                      metavar="FILE", help="input file")
    parser.add_option("-o", "--output", dest="output",
                      metavar="DIR", help="output dir")
    parser.add_option("-t", "--tmpdir", dest="tmpdir",
                      metavar="DIR", help="tmpdir dir")

    (options, args) = parser.parse_args()

    if not options.input or not options.output or not options.tmpdir:
        parser.print_help(sys.stderr)
        exit(1)

    if not os.path.exists(options.output):
        os.makedirs(options.output)

    try:
        try:
            xls_to_orders(options.input, options.output, options.tmpdir)
            db.session.commit()
        except Exception, inst:
            db.session.rollback()
            raise inst
    except Exception, inst:
        import traceback

        traceback.print_exc(sys.stderr)
        print >> sys.stderr, inst.message.encode('utf-8')
